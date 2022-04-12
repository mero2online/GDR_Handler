import openpyxl
import xlrd
import os
import numpy as np
from pandas import DataFrame

from HandleWord import HandleWord
from HelperFunc import resource_path, writeLocalFile


async def HandleExcel(excelFilename, input_file_extension):
    if (input_file_extension == '.xls'):
        workbook = xlrd.open_workbook(excelFilename)
        sh = workbook.sheet_by_index(workbook.nsheets-1)

        finalValues = []
        for rx in range(sh.nrows):
            finalValues.append(sh.row_values(rx))

        wb = openpyxl.Workbook()
        ws1 = wb.active

        for row in finalValues:
            ws1.append(row)

        wb.save(resource_path('input\\input.xlsx'))

        os.remove(resource_path('input\\input.xls'))
    else:
        workbook = openpyxl.load_workbook(excelFilename, data_only=True)

        worksheet = workbook.active

        all_rows = []
        for row in worksheet:
            current_row = []
            for cell in row:
                current_row.append(cell.value)
            all_rows.append(current_row)

        os.remove(resource_path('input\\input.xlsx'))

        wb = openpyxl.Workbook()
        ws1 = wb.active

        for row in all_rows:
            ws1.append(row)
            wb.save(resource_path('input\\input.xlsx'))

    handleInput()


def handleInput():
    workbook = openpyxl.load_workbook(resource_path('input\\input.xlsx'))

    ws1 = workbook.active

    df = DataFrame(ws1.values)
    rows = df.to_numpy().tolist()

    filtered = []
    filteredIdx = []
    filteredIdx.append(6)
    for r_idx, row in enumerate(rows):
        if r_idx == 6:
            filtered.append(row)
        elif r_idx >= 6:
            lastFiltered = filtered[len(filtered)-1]
            if [i for i, j in enumerate(row[3:len(row)-2]) if j != None] != [i for i, j in enumerate(lastFiltered[3:len(lastFiltered)-2]) if j != None]:
                filtered.append(row)
                filteredIdx.append(r_idx)

    filteredIdx.append(len(rows))

    secFiltered = []
    for idx, id in enumerate(filteredIdx):
        if idx < len(filteredIdx)-1:
            toIndex = filteredIdx[idx+1]
            secFiltered.append(rows[id:toIndex])

    rock = ["ANHYDRITE", "GYPSUM", "DOLOMITE", "DOLOMITIC LIMESTONE", "LIMESTONE", "MARL",
            "SHALE", "CLAY", "SILTSTONE", "SANDSTONE", "CEMENT"]

    allText = []
    for i, x in enumerate(secFiltered):
        new_arr = [[p if p != None else -999 for p in s]
                   for s in secFiltered[i]]
        new_arr_two = [[p if p != 'TR' else 0 for p in s] for s in new_arr]

        max_in_column = np.max(new_arr_two, axis=0)
        min_in_column = np.min(new_arr_two, axis=0)
        maxRockIndex = [index for index, value in sorted(
            enumerate(max_in_column[3:13]), reverse=True, key=lambda x: x[1]) if value > 1]

        maxRock = ['TR' if i == 0 else int(i) for i in max_in_column[3:13]]
        minRock = ['TR' if i == 0 else int(i) for i in min_in_column[3:13]]

        rocksPerInterval = []
        rockPercent = []
        for x in maxRockIndex:
            txtMinMax = f'({minRock[x]}-{maxRock[x]}%) {rock[x]}'
            txtMinEqMax = f'({maxRock[x]}%) {rock[x]}'
            txt = txtMinEqMax if minRock[x] == maxRock[x] else txtMinMax
            rockPercent.append(txt)
            rocksPerInterval.append(f'{rock[x]}:')

        allText.append(
            [f'FROM {int(min_in_column[0])} FT TO {int(max_in_column[1])} FT {", ".join(rockPercent)}:', rocksPerInterval])

    finalText = []
    for idx, x in enumerate(allText):
        finalText.append(x[0])
        for i in x[1]:
            finalText.append(i)

    writeLocalFile(resource_path(
        'output\\GEOLOGICAL_DESCRIPTION_REPORT.txt'), "\n".join(finalText))

    HandleWord(allText)
